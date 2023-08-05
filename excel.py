from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill

# Load the workbook
workbook = load_workbook(filename='Book1.xlsx')

# Select the worksheet
worksheet = workbook.active

color = None

for i in range (1, worksheet.max_row+1):
    row = worksheet[i]
    if(row[1].value is None):
        color = 'FF0000'
    else:
        if(row[2].value is None): color = "FFFFFF"
        elif('OPC' in worksheet[i][2].value): color = 'FFFF00'
        elif('SUP' in row[2].value): color = "FFBF8F"
        else: color = "FFFFFF"
        
    for j in range (worksheet.max_column):
        print(worksheet[i][j].value)
        worksheet[i][j].fill = PatternFill(fill_type='solid', start_color=color)
    print(color)
    print()

workbook.save('Book1.xlsx')
    